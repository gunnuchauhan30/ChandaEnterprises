"""
Import/export helpers for Material Master (and generic tabular reports).
Uses openpyxl directly -- no pandas dependency needed for this scope.
"""
import io
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from fastapi import UploadFile, HTTPException

MATERIAL_COLUMNS = [
    "material_code", "material_name", "category", "material_type", "grade", "size",
    "uom", "min_qty", "max_qty", "opening_qty", "unit_cost", "per_day_req",
    "warehouse", "rack", "bin", "hsn_code", "lead_time_days", "status",
]


def export_rows_to_excel(rows: List[Dict[str, Any]], columns: List[str], sheet_title: str = "Sheet1") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(col) + 2)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def parse_material_import_excel(file: UploadFile) -> List[Dict[str, Any]]:
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls files are supported for import")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    header = [str(c.value).strip().lower().replace(" ", "_") if c.value else "" for c in ws[1]]
    missing = {"material_code", "material_name"} - set(header)
    if missing:
        raise HTTPException(400, f"Import file missing required columns: {missing}")

    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(header, excel_row))
        if not record.get("material_code"):
            continue
        rows.append(record)
    return rows
